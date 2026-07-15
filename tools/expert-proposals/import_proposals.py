import os
import re
import json
import urllib.request
import urllib.error
import openpyxl
from pathlib import Path

SUPABASE_URL = 'https://qdqtkvyvhtjgxpxnvblk.supabase.co'
SUPABASE_KEY = 'sb_publishable_YXDm02hDBzLQmsUuVnZ_Og_IxQ60VCz'

SCRIPT_DIR = Path(__file__).resolve().parent
# Джерела даних перенесено з OneDrive у D:\pmg-data (15.07.2026); репо сайту — D:\rpe-pmg
PROJECT_ROOT = Path(r"D:\pmg-data")
EXCEL_FOLDER = PROJECT_ROOT / "16_робочі групи"
OUTPUT_JSON_PATH = SCRIPT_DIR / "data" / "expert_proposals.json"

def clean(text):
    if not text:
        return ""
    return re.sub(r"\s+", " ", str(text)).strip()

def parse_excel_files():
    proposals = []
    if not EXCEL_FOLDER.exists():
        print(f"[ПОМИЛКА] Папка {EXCEL_FOLDER} не існує!")
        return proposals

    files = sorted(list(EXCEL_FOLDER.glob("*.xlsx")))
    print(f"Знайдено {len(files)} файлів Excel для обробки.")

    for f_path in files:
        file_name = f_path.name
        if file_name.startswith('~$') or file_name.startswith('.'):
            continue

        print(f"Обробка файлу: {file_name}...")
        try:
            wb = openpyxl.load_workbook(f_path, data_only=True)
            sheet = wb.active
            
            # Спробуємо отримати номер та назву пакета з Рядка 2
            row2_val = ""
            for col in range(1, 10):
                val = sheet.cell(row=2, column=col).value
                if val:
                    row2_val = clean(val)
                    break
            
            package_number = None
            package_name = ""
            
            # Шукаємо "Пакет X «назва»"
            match = re.search(r"Пакет\s+(\d+)\s+«([^»]+)»", row2_val, re.IGNORECASE)
            if match:
                package_number = int(match.group(1))
                package_name = clean(match.group(2))
            else:
                # Спробуємо з першого рядка або імені файлу
                print(f"[ПОПЕРЕДЖЕННЯ] Не вдалося розпарсити номер пакета з рядка 2: '{row2_val}'")
                # Backup: імені файлу (наприклад "01_Інсульт.xlsx")
                file_num_match = re.search(r"^(\d+)", file_name)
                if file_num_match:
                    package_number = int(file_num_match.group(1))
                else:
                    package_number = 99
                package_name = file_name.replace(".xlsx", "")

            print(f"  -> Пакет {package_number}: {package_name[:60]}...")

            # Читаємо починаючи з рядка 5 (перші 4 рядки - це заголовки та опис)
            # Колонки: № з/п (1), Пакет / розділ (2), Пропозиція (3), Розбір (4), Позиція НСЗУ (5), Рішення (6)
            for row_idx in range(5, sheet.max_row + 1):
                num_val = sheet.cell(row=row_idx, column=1).value
                if num_val is None:
                    continue
                
                # Перевіримо, чи це дійсно число чи № з/п
                num_str = str(num_val).strip()
                if not num_str.isdigit():
                    continue
                
                row_num = int(num_str)
                item = clean(sheet.cell(row=row_idx, column=2).value)
                proposal_text = clean(sheet.cell(row=row_idx, column=3).value)
                analysis = clean(sheet.cell(row=row_idx, column=4).value)
                position_nhsu = clean(sheet.cell(row=row_idx, column=5).value)
                decision = clean(sheet.cell(row=row_idx, column=6).value)
                
                # Тільки якщо є хоча б пропозиція
                if proposal_text or item:
                    proposals.append({
                        "package_number": package_number,
                        "package_name": package_name,
                        "row_num": row_num,
                        "item": item,
                        "proposal": proposal_text,
                        "analysis": analysis,
                        "position_nhsu": position_nhsu,
                        "decision": decision,
                        "votes_clinical": [],
                        "votes_strategy": [],
                        "comments_clinical": [],
                        "comments_strategy": [],
                        "director_status": "pending",
                        "director_remarks": ""
                    })
        except Exception as e:
            print(f"[ПОМИЛКА] Помилка обробки файлу {file_name}: {e}")
            
    return proposals

def main():
    print("=======================================================")
    print("   Завантаження пропозицій експертів у Supabase   ")
    print("=======================================================")
    print()

    # 1. Scan and parse
    proposals = parse_excel_files()
    print(f"\nРазом розпарсено {len(proposals)} пропозицій експертів.")

    if not proposals:
        print("[ПОМИЛКА] Не знайдено або не вдалося розпарсити жодної пропозиції.")
        input("\nНатисніть Enter для виходу...")
        return

    # 2. Save static fallback JSON
    OUTPUT_JSON_PATH.parent.mkdir(parents=True, exist_ok=True)
    try:
        with open(OUTPUT_JSON_PATH, "w", encoding="utf-8") as f:
            json.dump({"proposals": proposals}, f, ensure_ascii=False, indent=2)
        print(f"[УСПІШНО] Локальну копію збережено у {OUTPUT_JSON_PATH.name}.")
    except Exception as e:
        print(f"[ПОПЕРЕДЖЕННЯ] Не вдалося зберегти локальний JSON: {e}")

    # 3. Get credentials
    print("\nВведіть ваші облікові дані на порталі (акаунт з роллю Admin/Director/Manager):")
    email = input("Email: ").strip()
    password = input("Пароль: ")

    # 4. Authenticate
    print("\nАвторизація в системі Supabase...")
    auth_url = f"{SUPABASE_URL}/auth/v1/token?grant_type=password"
    auth_payload = {
        "email": email,
        "password": password
    }
    headers = {
        "apikey": SUPABASE_KEY,
        "Content-Type": "application/json"
    }

    req = urllib.request.Request(
        auth_url,
        data=json.dumps(auth_payload).encode('utf-8'),
        headers=headers
    )
    
    try:
        with urllib.request.urlopen(req, timeout=10) as r:
            res_data = json.loads(r.read().decode('utf-8'))
            access_token = res_data["access_token"]
            user_data = res_data["user"]
            print(f"[УСПІШНО] Авторизовано як: {user_data.get('email')}")
    except urllib.error.HTTPError as e:
        body = e.read().decode('utf-8')
        print(f"[ПОМИЛКА] Не вдалося увійти в систему ({e.code}): {body}")
        input("\nНатисніть Enter для виходу...")
        return
    except Exception as e:
        print(f"[ПОМИЛКА] Помилка підключення: {e}")
        input("\nНатисніть Enter для виходу...")
        return

    auth_headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json"
    }

    # 5. Clear table in Supabase
    print("\nОчищення старої таблиці 'expert_proposals' у хмарі...")
    delete_url = f"{SUPABASE_URL}/rest/v1/expert_proposals?package_number=gt.0"
    del_req = urllib.request.Request(delete_url, method="DELETE", headers=auth_headers)
    
    try:
        with urllib.request.urlopen(del_req, timeout=15) as r:
            print("[УСПІШНО] Попередні записи видалено з бази.")
    except urllib.error.HTTPError as e:
        body = e.read().decode('utf-8')
        print(f"[ПОМИЛКА] Не вдалося очистити таблицю ({e.code}): {body}")
        print("\nПереконайтеся, що таблиця 'expert_proposals' була створена через SQL Editor.")
        input("\nНатисніть Enter для виходу...")
        return
    except Exception as e:
        print(f"[ПОМИЛКА] Не вдалося очистити таблицю: {e}")
        input("\nНатисніть Enter для виходу...")
        return

    # 6. Upload proposals to Supabase in chunks of 50
    chunk_size = 50
    total_proposals = len(proposals)
    print(f"\nЗавантаження {total_proposals} пропозицій пачками по {chunk_size}...")

    for i in range(0, total_proposals, chunk_size):
        chunk = proposals[i:i + chunk_size]
        end_idx = min(i + chunk_size, total_proposals)
        
        payload = []
        for p in chunk:
            payload.append({
                "package_number": int(p["package_number"]),
                "package_name": p["package_name"],
                "row_num": int(p["row_num"]),
                "item": p["item"],
                "proposal": p["proposal"],
                "analysis": p["analysis"],
                "position_nhsu": p["position_nhsu"],
                "decision": p["decision"],
                "votes_clinical": p["votes_clinical"],
                "votes_strategy": p["votes_strategy"],
                "comments_clinical": p.get("comments_clinical", []),
                "comments_strategy": p.get("comments_strategy", []),
                "director_status": p["director_status"],
                "director_remarks": p["director_remarks"]
            })

        print(f"Надсилання записів {i + 1} - {end_idx}...")
        
        upload_url = f"{SUPABASE_URL}/rest/v1/expert_proposals"
        upload_req = urllib.request.Request(
            upload_url,
            data=json.dumps(payload).encode('utf-8'),
            headers=auth_headers
        )
        
        try:
            with urllib.request.urlopen(upload_req, timeout=20) as r:
                pass
        except urllib.error.HTTPError as e:
            body = e.read().decode('utf-8')
            print(f"[ПОМИЛКА] Не вдалося завантажити записи {i+1}-{end_idx} ({e.code}): {body}")
            input("\nНатисніть Enter для виходу...")
            return
        except Exception as e:
            print(f"[ПОМИЛКА] Помилка завантаження записів: {e}")
            input("\nНатисніть Enter для виходу...")
            return

    print("\n=======================================================")
    print("   Синхронізація з базою Supabase успішно завершена!   ")
    print("=======================================================")
    input("\nНатисніть Enter для виходу...")

if __name__ == '__main__':
    main()

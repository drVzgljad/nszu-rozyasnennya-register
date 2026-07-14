import os
import json
import openpyxl
from pathlib import Path

# Paths
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)

XLSX_ICD_PATH = Path(r"c:\Users\Tartuga\OneDrive - National Health Service of Ukraine\!Работа!\!ПМГ!\класификаторы\ICD-AM формат.xlsx")
XLSX_ACHI_PATH = Path(r"c:\Users\Tartuga\OneDrive - National Health Service of Ukraine\!Работа!\!ПМГ!\класификаторы\klassifikator-intervencij.xlsx")

def build_icd_classifier():
    print(f"Loading ICD-AM from {XLSX_ICD_PATH}...")
    if not XLSX_ICD_PATH.exists():
        print(f"Error: ICD file not found at {XLSX_ICD_PATH}")
        return False
        
    wb = openpyxl.load_workbook(XLSX_ICD_PATH, read_only=True)
    sheet = wb["ICD-AM"]
    
    codes = {}
    for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if idx == 1:
            continue
            
        c4, n4 = row[4], row[5] # Disease code, Disease name
        c6, n6 = row[6], row[7] # Diagnosis code, Diagnosis name
        c8, n8 = row[8], row[9] # Refined diagnosis code, Refined name
        
        code, name = None, None
        if c8 and c8 not in ("_", "-", None) and str(c8).strip():
            code = str(c8).strip()
            name = str(n8).strip()
        elif c6 and c6 not in ("_", "-", None) and str(c6).strip():
            code = str(c6).strip()
            name = str(n6).strip()
        elif c4 and c4 not in ("_", "-", None) and str(c4).strip():
            code = str(c4).strip()
            name = str(n4).strip()
            
        if code and name:
            codes[code] = name
            
    out_path = DATA_DIR / "icd10_codes.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(codes, f, ensure_ascii=False, indent=2)
        
    print(f"ICD classifier compiled: {len(codes)} codes written to {out_path} ({out_path.stat().st_size} bytes)")
    return True

def build_achi_classifier():
    print(f"Loading ACHI from {XLSX_ACHI_PATH}...")
    if not XLSX_ACHI_PATH.exists():
        print(f"Error: ACHI file not found at {XLSX_ACHI_PATH}")
        return False
        
    wb = openpyxl.load_workbook(XLSX_ACHI_PATH, read_only=True)
    sheet = wb["кодиACHI"]
    
    codes = {}
    for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if idx < 28:
            continue
        name, code = row[1], row[2]
        if code and name:
            codes[str(code).strip()] = str(name).strip()
            
    out_path = DATA_DIR / "achi_codes.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(codes, f, ensure_ascii=False, indent=2)
        
    print(f"ACHI classifier compiled: {len(codes)} codes written to {out_path} ({out_path.stat().st_size} bytes)")
    return True

if __name__ == "__main__":
    icd_success = build_icd_classifier()
    achi_success = build_achi_classifier()
    if icd_success and achi_success:
        print("All classifiers compiled successfully!")
    else:
        print("Some classifiers failed to compile.")
